from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def df_to_styled_xlsx(df_ok, df_err, path: str):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Loaded OK"
    for r in dataframe_to_rows(df_ok, index=False, header=True):
        ws1.append(r)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell.alignment = Alignment(horizontal="center")
    for col in ws1.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 50)

    ws2 = wb.create_sheet("Errors")
    for r in dataframe_to_rows(df_err, index=False, header=True):
        ws2.append(r)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFC7CE")
        cell.alignment = Alignment(horizontal="center")
    for col in ws2.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 70)

    wb.save(path)
    return path
